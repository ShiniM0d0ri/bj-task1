import test_info
import openpyxl

wb = openpyxl.load_workbook('List of Recognized Startups.xlsx')
dirmail=[]

print("Enter starting row and ending row :D ")
rowinit=int(input())
rowend=int(input())


ws=wb.active
for i in range(rowinit,rowend+1):
    row=str(i)
    comp_name=ws['B'+row].value
    
    print("Company Name: ",comp_name)
    dirmail=test_info.spider(comp_name)

    ws['D'+row]=dirmail[0]
    ws['E'+row]=dirmail[1]
    print("Entered Successfully......")

wb.save("updated.xlsx")
print("updated.xlsx saved successfully")
print(":) :) :)")